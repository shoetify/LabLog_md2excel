import os
import openpyxl
from openpyxl.utils import get_column_letter

# Step 1: Ask for the file name
fileName = input("Please enter the file name: ")
excel_file_name = fileName + '.xlsx'
fileName += '.md'

# Step 2: Check if the file exists and read the file
if os.path.isfile(fileName):
    print(f"Reading '{fileName}' ...")

    # Step 3: Read the file
    try:
        with open(fileName, 'r') as file:
            content = file.read()
    except Exception as e:
        print(f"An error occurred while reading the file: {e}")
else:
    print(f"Error: The file '{fileName}' does not exist.")
    exit()

# Extracting table data
lines = content.split('\n')
tables = []
current_table = []
inside_table = False
table_count = 0
previous_line = ''

for line in lines:
    if "| --- " in line:
        if current_table:
            tables.append(current_table)
            current_table = []
        inside_table = True
        table_count += 1
        continue

    if inside_table:
        if not line.strip():
            if current_table:
                tables.append(current_table)
                current_table = []
            inside_table = False
            continue
        current_table.append(line.strip().split('|'))

if current_table:
    tables.append(current_table)

# Writing to Excel
wb = openpyxl.Workbook()

for idx, table in enumerate(tables, 1):
    ws = wb.create_sheet(title=f"Table {idx}")

    for row_idx, row in enumerate(table, 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value.strip())

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Remove the default sheet created automatically by openpyxl if no table found
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

wb.save(excel_file_name)
print(f"Table data has been written to {excel_file_name}")
